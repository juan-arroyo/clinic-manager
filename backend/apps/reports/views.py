# backend/apps/reports/views.py

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from apps.sales.models import Sale
from apps.users.models import Physio


def get_filtered_sales(request):
    """
    Filtra ventas según los parámetros de la URL (date_from, date_to, worker).
    Se reutiliza en la vista de reportes y en todas las exportaciones Excel.
    """
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    worker_id = request.GET.get('worker', '')

    sales = Sale.objects.select_related('patient', 'worker', 'bonus').order_by('-date')

    if date_from:
        sales = sales.filter(date__gte=date_from)  # gte = >= (mayor o igual)
    if date_to:
        sales = sales.filter(date__lte=date_to)    # lte = <= (menor o igual)
    if worker_id:
        sales = sales.filter(worker_id=worker_id)

    return sales


@login_required
def reports_view(request):
    """
    Página principal de reportes.
    Muestra filtros, resumen numérico y datos para los gráficos Chart.js.
    """
    sales = get_filtered_sales(request)

    sales_by_month = (
        sales
        .annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('month')
    )

    sales_by_worker = (
        sales
        .values('worker__first_name', 'worker__last_name')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )

    sales_by_treatment = (
        sales
        .values('treatment_type')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )

    # Chart.js necesita listas separadas de etiquetas y valores
    MONTHS_ES = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }

    months_labels = [
        f"{MONTHS_ES[s['month'].month]} {s['month'].year}" if s['month'] else ''
        for s in sales_by_month
    ]
    months_data = [float(s['total']) for s in sales_by_month]
    worker_labels = [f"{s['worker__first_name']}" for s in sales_by_worker]
    worker_data = [float(s['total']) for s in sales_by_worker]
    treatment_labels = [s['treatment_type'] for s in sales_by_treatment]
    treatment_data = [float(s['total']) for s in sales_by_treatment]

    total_amount = sales.aggregate(total=Sum('amount'))['total'] or 0
    total_sales = sales.count()
    workers = Physio.objects.filter(is_active=True).order_by('first_name')

    return render(request, 'reports/index.html', {
        'sales': sales,
        'workers': workers,
        'total_amount': total_amount,
        'total_sales': total_sales,
        # Listas para Chart.js — se convierten a JSON en el template con |safe
        'months_labels': months_labels,
        'months_data': months_data,
        'worker_labels': worker_labels,
        'worker_data': worker_data,
        'treatment_labels': treatment_labels,
        'treatment_data': treatment_data,
        # Filtros activos — se devuelven para mantener los valores en el formulario
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'selected_worker': request.GET.get('worker', ''),
    })


def build_excel_styles():
    """
    Devuelve los estilos comunes para cabeceras Excel.
    Se define una sola vez y se reutiliza en los tres exports.
    """
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2596BE', end_color='2596BE', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    return header_font, header_fill, header_alignment


def apply_header_style(ws, headers):
    """
    Aplica estilo visual a la fila de cabeceras de una hoja Excel.
    """
    header_font, header_fill, header_alignment = build_excel_styles()
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


@login_required
def export_sales_excel(request):
    """Exporta las ventas filtradas a Excel."""
    sales = get_filtered_sales(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    # Columnas acordadas con la clínica — no cambiar el orden sin revisar el proceso contable
    headers = [
        'Factura emitida', 'Fecha', 'Nombre', 'Importe',
        'Bono', 'Método de pago', 'Trabajador', 'Descripción'
    ]
    apply_header_style(ws, headers)

    for sale in sales:
        ws.append([
            'Sí' if sale.invoice_issued else 'No',
            sale.date.strftime('%d/%m/%Y'),
            sale.patient.full_name,
            float(sale.amount),
            sale.bonus.bonus_number if sale.bonus else '',
            sale.get_payment_method_display(),
            sale.worker.full_name,
            sale.description,
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ventas.xlsx"'
    wb.save(response)
    return response


@login_required
def export_invoices_excel(request):
    """Exporta solo las ventas con factura emitida a Excel."""
    sales = get_filtered_sales(request).filter(invoice_issued=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Facturas'

    headers = [
        'Número de factura', 'Fecha', 'Empresa/Cliente',
        'Método de pago', 'Importe', 'Observación'
    ]
    apply_header_style(ws, headers)

    for sale in sales:
        ws.append([
            sale.invoice_number or '',
            sale.date.strftime('%d/%m/%Y'),
            sale.patient.full_name,
            sale.get_payment_method_display(),
            float(sale.amount),
            sale.description,
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="facturas.xlsx"'
    wb.save(response)
    return response


@login_required
def export_fisio_excel(request):
    """Exporta las ventas filtradas por fisioterapeuta a Excel."""
    sales = get_filtered_sales(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas por fisio'

    headers = [
        'Factura emitida', 'Fecha', 'Nombre', 'Importe',
        'Bono', 'Método de pago', 'Trabajador', 'Descripción'
    ]
    apply_header_style(ws, headers)

    for sale in sales:
        ws.append([
            'Sí' if sale.invoice_issued else 'No',
            sale.date.strftime('%d/%m/%Y'),
            sale.patient.full_name,
            float(sale.amount),
            sale.bonus.bonus_number if sale.bonus else '',
            sale.get_payment_method_display(),
            sale.worker.full_name,
            sale.description,
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ventas-fisio.xlsx"'
    wb.save(response)
    return response