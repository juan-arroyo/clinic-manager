# backend/apps/reports/tests.py

from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model
import openpyxl
import io

from apps.users.models import Physio
from apps.patients.models import Patient
from apps.sales.models import Sale

User = get_user_model()


class ExcelVentasTest(TestCase):

    def setUp(self):
        """Se ejecuta antes de cada test. Prepara todos los datos necesarios."""

        # Tu modelo User usa EMAIL para login, no username
        # Por eso creamos el usuario con email y hacemos login con email
        self.user = User.objects.create_user(
            username='test_user',
            email='test@clinicasolaz.com',   # campo obligatorio y único
            password='test_password_123',
            first_name='Test',               # REQUIRED_FIELDS lo exige
            last_name='Usuario',
        )

        self.client = Client()

        # Login con EMAIL, no con username — así funciona tu sistema
        self.client.login(
            username='test@clinicasolaz.com',  # Django usa USERNAME_FIELD aquí
            password='test_password_123'
        )

        self.fisio = Physio.objects.create(
            first_name='David',
            last_name='Pachone',
            color='#2596be',
            is_active=True
        )

        self.patient = Patient.objects.create(
            dni='12345678A',
            first_name='Marta',
            last_name='Gil García',
            phone='600000001'
        )

        self.sale = Sale.objects.create(
            patient=self.patient,
            worker=self.fisio,
            amount=35,
            payment_method='efectivo',
            treatment_type='general',
            invoice_issued=False,
        )

    def test_excel_ventas_cabeceras_correctas(self):
        """Verifica que el Excel de ventas tiene las columnas correctas."""

        response = self.client.get(reverse('reports:export_sales'))

        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])

        excel_file = io.BytesIO(response.content)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        primera_fila = [cell.value for cell in ws[1]]

        cabeceras_esperadas = [
            'Factura emitida', 'Fecha', 'Nombre', 'Importe',
            'Bono', 'Método de pago', 'Trabajador', 'Descripción'
        ]

        self.assertEqual(primera_fila, cabeceras_esperadas)


    def test_excel_ventas_datos_correctos(self):
        """Verifica que los datos de la primera venta se exportan correctamente."""

        response = self.client.get(reverse('reports:export_sales'))
        excel_file = io.BytesIO(response.content)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # La fila 1 son cabeceras, la fila 2 es la primera venta
        segunda_fila = [cell.value for cell in ws[2]]

        # Verificamos cada columna por separado para saber exactamente cuál falla
        self.assertEqual(segunda_fila[0], 'No')               # Factura emitida
        self.assertEqual(segunda_fila[2], 'Marta Gil García') # Nombre
        self.assertEqual(segunda_fila[3], 35.0)               # Importe
        self.assertIsNone(segunda_fila[4])                     # Bono vacío → None en Excel
        self.assertEqual(segunda_fila[5], 'Efectivo')          # Método de pago
        self.assertEqual(segunda_fila[6], 'David Pachone')     # Trabajador

    def test_excel_ventas_descarga_con_nombre_correcto(self):
        """Verifica que el archivo se descarga con el nombre ventas.xlsx"""

        response = self.client.get(reverse('reports:export_sales'))

        # Content-Disposition es la cabecera HTTP que dice cómo se llama el archivo
        self.assertIn(
            'ventas.xlsx',
            response['Content-Disposition']
        )


class ExcelFacturasTest(TestCase):
    """
    Tests para la exportación Excel de facturas.
    """

    def setUp(self):
        """Prepara datos con una venta QUE SÍ tiene factura emitida."""

        self.user = User.objects.create_user(
            username='test_user2',
            email='test2@clinicasolaz.com',
            password='test_password_123',
            first_name='Test',
            last_name='Usuario',
        )
        self.client = Client()
        self.client.login(
            username='test2@clinicasolaz.com',
            password='test_password_123'
        )

        self.fisio = Physio.objects.create(
            first_name='Matias',
            last_name='Fermu',
            color='#ff5733',
            is_active=True
        )

        self.patient = Patient.objects.create(
            dni='87654321B',
            first_name='Pedro',
            last_name='Gutiérrez Gómez',
        )

        # Esta venta SÍ tiene factura emitida — aparecerá en el Excel de facturas
        self.sale = Sale.objects.create(
            patient=self.patient,
            worker=self.fisio,
            amount=35,
            payment_method='efectivo',
            treatment_type='general',
            invoice_issued=True,   # ← clave: tiene factura
        )

    def test_excel_facturas_cabeceras_correctas(self):
        """Verifica que el Excel de facturas tiene las columnas correctas."""

        response = self.client.get(reverse('reports:export_invoices'))

        self.assertEqual(response.status_code, 200)

        excel_file = io.BytesIO(response.content)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        primera_fila = [cell.value for cell in ws[1]]

        cabeceras_esperadas = [
            'Número de factura', 'Fecha', 'Empresa/Cliente',
            'Método de pago', 'Importe', 'Observación'
        ]

        self.assertEqual(primera_fila, cabeceras_esperadas)

    def test_excel_facturas_solo_incluye_facturas_emitidas(self):
        """
        Verifica que el Excel de facturas SOLO contiene ventas con factura emitida.
        Una venta sin factura no debe aparecer.
        """

        # Creamos una segunda venta SIN factura — no debe salir en el Excel
        Sale.objects.create(
            patient=self.patient,
            worker=self.fisio,
            amount=50,
            payment_method='tarjeta',
            treatment_type='general',
            invoice_issued=False,   # ← sin factura
        )

        response = self.client.get(reverse('reports:export_invoices'))
        excel_file = io.BytesIO(response.content)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # La fila 1 son cabeceras, así que el número de filas de datos es max_row - 1
        filas_de_datos = ws.max_row - 1

        # Solo debe haber 1 fila — la venta con factura, no la de 50€ sin factura
        self.assertEqual(filas_de_datos, 1)

    def test_excel_facturas_nombre_archivo_correcto(self):
        """Verifica que el archivo se descarga con el nombre facturas.xlsx"""

        response = self.client.get(reverse('reports:export_invoices'))
        self.assertIn('facturas.xlsx', response['Content-Disposition'])


class ExcelFisioTest(TestCase):
    """
    Tests para la exportación Excel de ventas por fisioterapeuta.
    """

    def setUp(self):

        self.user = User.objects.create_user(
            username='test_user3',
            email='test3@clinicasolaz.com',
            password='test_password_123',
            first_name='Test',
            last_name='Usuario',
        )
        self.client = Client()
        self.client.login(
            username='test3@clinicasolaz.com',
            password='test_password_123'
        )

        # Creamos DOS fisios para verificar el filtrado
        self.fisio_david = Physio.objects.create(
            first_name='David',
            last_name='Pachone',
            color='#2596be',
            is_active=True
        )
        self.fisio_matias = Physio.objects.create(
            first_name='Matias',
            last_name='Fermu',
            color='#ff5733',
            is_active=True
        )

        self.patient = Patient.objects.create(
            dni='11111111C',
            first_name='Laura',
            last_name='Blanco Serrano',
        )

        # Una venta de David
        Sale.objects.create(
            patient=self.patient,
            worker=self.fisio_david,
            amount=60,
            payment_method='efectivo',
            treatment_type='general',
            invoice_issued=False,
        )

        # Una venta de Matias
        Sale.objects.create(
            patient=self.patient,
            worker=self.fisio_matias,
            amount=45,
            payment_method='tarjeta',
            treatment_type='general',
            invoice_issued=False,
        )

    def test_excel_fisio_cabeceras_correctas(self):
        """Verifica que el Excel de ventas por fisio tiene las columnas correctas."""

        response = self.client.get(reverse('reports:export_fisio'))

        self.assertEqual(response.status_code, 200)

        excel_file = io.BytesIO(response.content)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        primera_fila = [cell.value for cell in ws[1]]

        cabeceras_esperadas = [
            'Factura emitida', 'Fecha', 'Nombre', 'Importe',
            'Bono', 'Método de pago', 'Trabajador', 'Descripción'
        ]

        self.assertEqual(primera_fila, cabeceras_esperadas)

    def test_excel_fisio_filtra_por_trabajador(self):
        """
        Verifica que al filtrar por fisio solo aparecen sus ventas.
        Si filtramos por David, no debe aparecer la venta de Matias.
        """

        # Filtramos por el ID de David pasándolo como parámetro en la URL
        url = reverse('reports:export_fisio') + f'?worker={self.fisio_david.pk}'
        response = self.client.get(url)

        excel_file = io.BytesIO(response.content)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        filas_de_datos = ws.max_row - 1

        # Solo debe aparecer la venta de David (60€), no la de Matias (45€)
        self.assertEqual(filas_de_datos, 1)

        # Confirmamos que el trabajador de esa fila es David
        trabajador_en_excel = ws.cell(row=2, column=7).value
        self.assertEqual(trabajador_en_excel, 'David Pachone')